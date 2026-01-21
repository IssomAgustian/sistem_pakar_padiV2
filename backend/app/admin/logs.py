"""Admin - Activity Logs"""
from flask import Blueprint, jsonify, request, render_template, send_file
from flask_jwt_extended import jwt_required
from datetime import datetime, timedelta
from app.models.admin_log import AdminLog
from app.models.user import User
from app.utils.decorators import admin_required
from io import BytesIO
import csv
from openpyxl import Workbook
from reportlab.lib.pagesizes import letter, A4
from reportlab.lib import colors
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet

bp = Blueprint('admin_logs', __name__)

@bp.route('/', methods=['GET'])
def logs_page():
    """Render logs page"""
    return render_template('admin/logs.html')

@bp.route('/api/logs', methods=['GET'])
def get_logs_api():
    """Get all logs for frontend (without JWT)"""
    try:
        # Get all logs with admin info
        logs = AdminLog.query.join(User, AdminLog.admin_id == User.id)\
            .add_columns(
                AdminLog.id,
                AdminLog.action,
                AdminLog.table_name,
                AdminLog.description,
                AdminLog.ip_address,
                AdminLog.created_at,
                User.full_name.label('admin_name')
            ).order_by(AdminLog.created_at.desc()).all()

        logs_data = []
        for log in logs:
            logs_data.append({
                'id': log.id,
                'admin_name': log.admin_name,
                'action': log.action,
                'table_name': log.table_name,
                'details': log.description,
                'ip_address': log.ip_address,
                'created_at': log.created_at.isoformat() if log.created_at else None
            })

        return jsonify({
            'success': True,
            'logs': logs_data
        })
    except Exception as e:
        return jsonify({
            'success': False,
            'message': str(e),
            'logs': []
        }), 500

@bp.route('/list', methods=['GET'])
@jwt_required()
@admin_required
def get_all_logs():
    """Get all admin activity logs with pagination"""
    page = int(request.args.get('page', 1))
    per_page = int(request.args.get('per_page', 50))
    admin_id = request.args.get('admin_id', type=int)
    action = request.args.get('action', '')
    date_from = request.args.get('date_from', '')
    date_to = request.args.get('date_to', '')

    query = AdminLog.query

    if admin_id:
        query = query.filter_by(admin_id=admin_id)

    if action:
        query = query.filter_by(action=action)

    if date_from:
        try:
            from_date = datetime.strptime(date_from, '%Y-%m-%d')
            query = query.filter(AdminLog.created_at >= from_date)
        except ValueError:
            pass

    if date_to:
        try:
            to_date = datetime.strptime(date_to, '%Y-%m-%d')
            to_date = to_date.replace(hour=23, minute=59, second=59)
            query = query.filter(AdminLog.created_at <= to_date)
        except ValueError:
            pass

    pagination = query.order_by(AdminLog.created_at.desc()).paginate(
        page=page, per_page=per_page, error_out=False
    )

    return jsonify({
        'success': True,
        'data': [log.to_dict() for log in pagination.items],
        'pagination': {
            'page': page,
            'per_page': per_page,
            'total': pagination.total,
            'pages': pagination.pages
        }
    })

@bp.route('/stats', methods=['GET'])
@jwt_required()
@admin_required
def get_log_stats():
    """Get activity log statistics"""
    from sqlalchemy import func

    # Total logs
    total_logs = AdminLog.query.count()

    # Logs today
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    logs_today = AdminLog.query.filter(AdminLog.created_at >= today_start).count()

    # Logs this week
    week_start = datetime.utcnow() - timedelta(days=7)
    logs_week = AdminLog.query.filter(AdminLog.created_at >= week_start).count()

    # Most common actions
    common_actions = AdminLog.query.with_entities(
        AdminLog.action,
        func.count(AdminLog.id).label('count')
    ).group_by(AdminLog.action).order_by(
        func.count(AdminLog.id).desc()
    ).limit(10).all()

    # Most active admins
    active_admins = AdminLog.query.join(User).with_entities(
        User.email,
        User.full_name,
        func.count(AdminLog.id).label('activity_count')
    ).group_by(User.email, User.full_name).order_by(
        func.count(AdminLog.id).desc()
    ).limit(5).all()

    return jsonify({
        'success': True,
        'data': {
            'total_logs': total_logs,
            'logs_today': logs_today,
            'logs_week': logs_week,
            'common_actions': [{'action': a[0], 'count': a[1]} for a in common_actions],
            'active_admins': [{'email': a[0], 'full_name': a[1], 'count': a[2]} for a in active_admins]
        }
    })

@bp.route('/actions', methods=['GET'])
@jwt_required()
@admin_required
def get_action_types():
    """Get all distinct action types"""
    from sqlalchemy import distinct

    actions = AdminLog.query.with_entities(
        distinct(AdminLog.action)
    ).order_by(AdminLog.action).all()

    return jsonify({
        'success': True,
        'data': [a[0] for a in actions]
    })

@bp.route('/api/logs/export', methods=['POST'])
def export_logs():
    """Export logs to Excel or PDF"""
    try:
        data = request.get_json()
        export_format = request.args.get('format', 'excel')
        start_date = data.get('start_date', '')
        end_date = data.get('end_date', '')
        action = data.get('action', '')
        table = data.get('table', '')

        # Build query
        query = AdminLog.query.join(User, AdminLog.admin_id == User.id)\
            .add_columns(
                AdminLog.id,
                AdminLog.action,
                AdminLog.table_name,
                AdminLog.description,
                AdminLog.ip_address,
                AdminLog.created_at,
                User.full_name.label('admin_name')
            )

        # Apply filters
        if start_date:
            try:
                from_date = datetime.strptime(start_date, '%Y-%m-%d')
                query = query.filter(AdminLog.created_at >= from_date)
            except ValueError:
                pass

        if end_date:
            try:
                to_date = datetime.strptime(end_date, '%Y-%m-%d')
                to_date = to_date.replace(hour=23, minute=59, second=59)
                query = query.filter(AdminLog.created_at <= to_date)
            except ValueError:
                pass

        if action:
            query = query.filter(AdminLog.action == action)

        if table:
            query = query.filter(AdminLog.table_name == table)

        logs = query.order_by(AdminLog.created_at.desc()).all()

        if export_format == 'excel':
            return export_to_excel(logs)
        elif export_format == 'pdf':
            return export_to_pdf(logs)
        else:
            return jsonify({'success': False, 'message': 'Invalid format'}), 400

    except Exception as e:
        return jsonify({'success': False, 'message': str(e)}), 500

def export_to_excel(logs):
    """Export logs to Excel format"""
    wb = Workbook()
    ws = wb.active
    ws.title = "System Logs"

    # Headers
    headers = ['No', 'Admin', 'Action', 'Table', 'Details', 'IP Address', 'Timestamp']
    ws.append(headers)

    # Style headers
    from openpyxl.styles import Font, PatternFill
    for cell in ws[1]:
        cell.font = Font(bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')

    # Data
    for idx, log in enumerate(logs, start=1):
        ws.append([
            idx,
            log.admin_name,
            log.action,
            log.table_name or '-',
            log.description or '-',
            log.ip_address or '-',
            log.created_at.strftime('%Y-%m-%d %H:%M:%S') if log.created_at else '-'
        ])

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width

    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=f'system_logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
    )

def export_to_pdf(logs):
    """Export logs to PDF format"""
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    elements = []

    # Styles
    styles = getSampleStyleSheet()
    title_style = styles['Heading1']

    # Title
    title = Paragraph("System Logs Report", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.3*inch))

    # Metadata
    metadata = Paragraph(
        f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}<br/>"
        f"Total Records: {len(logs)}",
        styles['Normal']
    )
    elements.append(metadata)
    elements.append(Spacer(1, 0.3*inch))

    # Table data
    table_data = [['No', 'Admin', 'Action', 'Table', 'Timestamp']]

    for idx, log in enumerate(logs, start=1):
        table_data.append([
            str(idx),
            log.admin_name[:20] if log.admin_name else '-',
            log.action,
            log.table_name or '-',
            log.created_at.strftime('%Y-%m-%d %H:%M') if log.created_at else '-'
        ])

    # Create table
    table = Table(table_data, colWidths=[0.5*inch, 1.5*inch, 1*inch, 1*inch, 1.5*inch])

    # Style table
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#4472C4')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.lightgrey]),
    ]))

    elements.append(table)

    # Build PDF
    doc.build(elements)
    buffer.seek(0)

    return send_file(
        buffer,
        mimetype='application/pdf',
        as_attachment=True,
        download_name=f'system_logs_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf'
    )
